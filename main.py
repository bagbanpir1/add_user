import openpyxl
import requests
from concurrent.futures import ProcessPoolExecutor


def create_api(payload):
    try:
        api_url = '' 
        response = requests.post(api_url, json=payload)

        if response.status_code == 201:
            print(f"Successfully created api for id {payload['id']}")
        else:
            print(f"Failed to create api for id {payload['id']}. Status code: {response.status_code}")
    except Exception as e:
        print(f"Error for id {payload['id']}: {str(e)}")


def parsing(sheet):
    lst = []
    for row in sheet.iter_rows(min_row=10, max_row=sheet.max_row,
                               min_col=1, max_col=sheet.max_column):
        lst.append([cell.value for cell in row])

    lst = [i for i in lst if type(i[0]) == type(1)]
    return lst


def main():
    excel_file_path = 'file.xlsx'
    sheet_name = 'Sheet1'
    
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[sheet_name]
    lst = parsing(sheet)
    
    num_processes = 4
    payloads = []
    
    for row in lst:
        full_name = list(map(str, row[2].split()))
        last_name = full_name[0]
        first_name = full_name[1]
        middle_name = full_name[2]
        username = first_name + "_" + last_name
        email = last_name + first_name + "@gmail.com"
        position = row[1]
        age = row[4]
        
        payload = {
                'last_name':last_name,
                'first_name':first_name,
                'middle_name':middle_name,
                'username':username,
                'position': position,
                'email':email,
                'age':age,
                } 
        payloads.append(payload)
        
    print(payloads, end="\n")
    print()
    
    with ProcessPoolExecutor(max_workers=num_processes) as executor:
        executor.map(create_api, payloads)


if __name__ == "__main__":
    main()
